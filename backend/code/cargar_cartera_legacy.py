"""Carga la deuda anterior a jun-2025 desde la exportacion a Excel del reporte de
cartera de Siigo, RESTANDO lo que ya trae la API (para no contar doble).

La API de Siigo no expone facturas anteriores al 2025-06-03, asi que esa deuda
vieja hay que traerla de la exportacion manual. Ver backend/supabase_cartera_legacy.sql.

Uso: python cargar_cartera_legacy.py "<ruta del xlsx>"
"""
import urllib.request, urllib.error, json, time, sys
from collections import defaultdict
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\johan\Downloads\Cuentas por cobrar general por cliente (1).xlsx'
SB = 'https://wavqyesyqqmawjfaztvb.supabase.co'
SR = '__SUPABASE_SERVICE_ROLE__'
H = {'apikey': SR, 'Authorization': 'Bearer ' + SR, 'Content-Type': 'application/json'}


def req(url, data=None, method=None, tries=8):
    for i in range(tries):
        try:
            return urllib.request.urlopen(
                urllib.request.Request(url, data=data, headers=H, method=method), timeout=60).read()
        except urllib.error.HTTPError as e:
            return b'HTTPERR:' + str(e.code).encode() + b' ' + e.read()[:300]
        except Exception:
            time.sleep(4)
    raise RuntimeError('sin red')


# 1) Reporte de Siigo
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.worksheets[0]
siigo, corte = {}, None
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if not a:
        continue
    s = str(a).strip()
    if s.startswith('Procesado en:'):
        corte = s.replace('Procesado en:', '').strip()
    if not s[:1].isdigit():
        continue
    siigo[s] = {'cliente': ws.cell(r, 2).value, 'total': float(ws.cell(r, 10).value or 0)}
print(f'reporte Siigo: {len(siigo)} clientes | corte: {corte}')
print(f'  total reporte: {sum(v["total"] for v in siigo.values()):,.2f}')

# 2) Cartera viva (la que si trae la API)
api = defaultdict(float)
rows = json.loads(req(SB + '/rest/v1/cartera?select=nit,saldo_cop&limit=5000'))
for x in rows:
    api[str(x['nit']).strip()] += float(x['saldo_cop'] or 0)
print(f'  total API    : {sum(api.values()):,.2f}')

# 3) legacy = Siigo - API (por cliente)
legacy = []
for nit in sorted(set(list(siigo) + list(api))):
    s = siigo.get(nit, {}).get('total', 0.0)
    a = api.get(nit, 0.0)
    dif = round(s - a, 2)
    if abs(dif) < 1:
        continue
    legacy.append({
        'nit': nit,
        'cliente': siigo.get(nit, {}).get('cliente') or '(solo en la API)',
        'saldo_cop': dif,
        # Todo lo legacy es anterior a jun-2025 => mas de 91 dias vencido.
        'bucket': 'saldo_favor' if dif < 0 else 'v91',
        'fecha_corte': time.strftime('%Y-%m-%d'),
        'nota': 'Deuda anterior a jun-2025: la API de Siigo no expone facturas previas al 2025-06-03. Cargado desde la exportacion del reporte de cartera.'
    })

print(f'\nclientes con deuda vieja: {len(legacy)}')
for l in legacy:
    print(f"   {l['nit']:>12} {str(l['cliente'])[:36]:<36} {l['saldo_cop']:>15,.2f}  [{l['bucket']}]")
tot_leg = sum(l['saldo_cop'] for l in legacy)
print(f'\n   TOTAL legacy: {tot_leg:,.2f}')
print(f'   API + legacy: {sum(api.values()) + tot_leg:,.2f}')
print(f'   Siigo dice  : {sum(v["total"] for v in siigo.values()):,.2f}')
if abs((sum(api.values()) + tot_leg) - sum(v['total'] for v in siigo.values())) > 1:
    print('   *** NO CUADRA — no se carga nada ***')
    raise SystemExit(1)
print('   -> cuadra')

# 4) Reemplazar la tabla
req(SB + '/rest/v1/cartera_legacy?nit=neq.__ninguno__', method='DELETE')
out = req(SB + '/rest/v1/cartera_legacy', data=json.dumps(legacy).encode(), method='POST')
print('\ninsert:', out[:120] if out else 'OK')
chk = json.loads(req(SB + '/rest/v1/rpc/cartera_resumen_total', data=b'{}', method='POST'))
print('RPC cartera_resumen_total ->', json.dumps(chk, ensure_ascii=False))
